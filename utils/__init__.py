from random import uniform
import json
from time import sleep
import time
from pathlib import Path
from os.path import join
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from ua_generator.options import Options as OptionsUA
from ua_generator.data.version import VersionRange
import openpyxl
import ua_generator


def get_random_user_agent() -> dict:
    options = OptionsUA()
    options.version_ranges = {
        'chrome': VersionRange(140, 144),  # Choose version between 125 and 129
    }
    ua = ua_generator.generate(
        browser='chrome', platform='windows', options=options)
    ua.headers.accept_ch(
        "Sec-CH-UA-Platform-Version, Sec-CH-UA-Full-Version-List")
    # return ua.headers.get()
    headers = ua.headers.get()
    return {k.title(): v for k, v in headers.items()}


def delay(min: float, max: float):
    sleep(uniform(min, max))


def write_json(filename: str, data: list[dict]) -> None:
    with open(filename, "w+") as f:
        json.dump(data, f, indent=4)


def setup_driver(headless: bool = False, pref: dict = {}) -> WebDriver:
    ua = get_random_user_agent()

    options = Options()
    if headless:
        options.add_argument("--headless=new")  # New headless mode

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1920,1080")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("prefs", pref)
    options.add_argument("--disable-images")
    options.add_argument(f"--user-agent={ua.get("User-Agent")}")
    options.add_argument("--blink-settings=imagesEnabled=false")
    return webdriver.Chrome(options=options)


def find_element_or_none(wait: WebDriverWait, selector: str) -> WebElement | None:
    try:
        elm = wait.until(EC.presence_of_element_located((By.XPATH, selector)))
        return elm
    except:
        return None


def find_elements(wait: WebDriverWait, selector: str) -> list[WebElement] | None:
    try:
        children = wait.until(
            EC.visibility_of_any_elements_located((By.XPATH, selector))
        )
        return children
    except:
        return None


def clean_spreadsheet(filename: str) -> None:
    wb = openpyxl.load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.value = None
                cell.hyperlink = None
    wb.save(filename)
    wb.close()


def get_xlsx_filepath(filename: str) -> str:
    project_root = Path(__file__).resolve().parent.parent
    return join(project_root, "spreadsheet", filename)


def get_with_backoff(driver: WebDriver, url: str, max_retries=5, initial_delay=2):
    retries = 0
    delay = initial_delay

    while retries < max_retries:
        try:
            driver.get(url)
            return True  # Exit function on success

        except TimeoutException:
            retries += 1
            if retries == max_retries:
                print("Max retries reached. Failing...")
                raise

            print(f"Timeout! Retrying [{retries+1}] in {delay} seconds...")
            time.sleep(delay)
            delay *= 2  # Exponentially increase the wait (2, 4, 8, 16...)


def save_xlsx(xlsx_path: str, funds: list[dict], cols: list[str], sheet: str, start: int = 2):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet]
    for fund in funds:
        for idx, val in enumerate(cols):
            col = idx+1
            row = fund.get("index")
            if row:
                start = row
            if val == "url":
                cell = ws.cell(start, col, fund.get(val))
                cell.style = "Hyperlink"
                cell.hyperlink = fund.get(val)
                continue
            ws.cell(start, col, fund.get(val))
        start += 1
    wb.save(xlsx_path)
    wb.close()
