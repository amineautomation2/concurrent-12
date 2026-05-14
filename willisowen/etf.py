import curl_cffi
import openpyxl
from math import ceil
from utils import delay


def Willis_ETF(h: dict, out_xlsx: str) -> None:
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb["ETF"]
    sheet_iter = 2

    api_endpoint = 'https://lt.morningstar.com/api/rest.svc/8c7eme98am/security/screener?page=1&pageSize=10&sortOrder=LegalName%20asc&outputType=json&version=1&languageId=en-GB&currencyId=GBP&universeIds=FCGBR%24%24ALL_4588&securityDataPoints=SecId%7CName%7CTenforeId%7CLegalName%7CIsin%7CBaseCurrency%7CClosePrice%7CPriceCurrency%7CAnalystRatingScale%7CStarRatingM255%7CSustainabilityRank%7CLargestSector%7CYield_M12%7CGBRReturnM0%7CGBRReturnM12%7CGBRReturnM36%7CGBRReturnM60%7CGBRReturnM120%7CYR_ReturnM12_1%7CYR_ReturnM12_2%7CYR_ReturnM12_3%7CYR_ReturnM12_4%7CYR_ReturnM12_5%7COngoingCostEstimated%7CPerformanceFeeEstimated%7CTransactionFeeEstimated%7CTrackRecordExtension&filters=&term=&subUniverseId='

    headers = {
        'authority': 'lt.morningstar.com',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'en-US,en;q=0.9',
        'origin': 'https://www.willisowen.co.uk',
        'referer': 'https://www.willisowen.co.uk/',
    }
    headers.update(h)
    response = curl_cffi.get(
        api_endpoint, headers=headers, impersonate="chrome")
    api_obj = response.json()
    total_funds = int(api_obj["total"])
    page_size = 50
    total_pages = ceil(total_funds / page_size)

    print(f'[Willis Owen] Total ETF = {total_funds}')

    for page in range(1, total_pages + 1):
        url = f'https://lt.morningstar.com/api/rest.svc/8c7eme98am/security/screener?page={page}&pageSize={page_size}&sortOrder=LegalName%20asc&outputType=json&version=1&languageId=en-GB&currencyId=GBP&universeIds=FCGBR%24%24ALL_4588&securityDataPoints=SecId%7CName%7CTenforeId%7CLegalName%7CIsin%7CBaseCurrency%7CClosePrice%7CPriceCurrency%7CAnalystRatingScale%7CStarRatingM255%7CSustainabilityRank%7CLargestSector%7CYield_M12%7CGBRReturnM0%7CGBRReturnM12%7CGBRReturnM36%7CGBRReturnM60%7CGBRReturnM120%7CYR_ReturnM12_1%7CYR_ReturnM12_2%7CYR_ReturnM12_3%7CYR_ReturnM12_4%7CYR_ReturnM12_5%7COngoingCostEstimated%7CPerformanceFeeEstimated%7CTransactionFeeEstimated%7CTrackRecordExtension&filters=&term=&subUniverseId='
        print(f'[#] ETF Page {page}/{total_pages} [#]')

        response = curl_cffi.get(url, headers=headers, impersonate="chrome")
        list_funds = response.json()["rows"]

        for fund in list_funds:
            isin = fund["Isin"]
            ws.cell(sheet_iter, 1).value = fund['LegalName']
            ws.cell(sheet_iter, 2).value = isin
            fund_url = f"https://www.willisowen.co.uk/explore/search-fund-detail?ISIN={isin}#?id={isin}&idCurrencyId=%20&idType=ISIN&marketCode=%20"
            c = ws.cell(sheet_iter, 3, fund_url)
            c.hyperlink = fund_url
            c.style = "Hyperlink"
            sheet_iter += 1

        wb.save(out_xlsx)
        delay(1, 2)
    wb.save(out_xlsx)
    wb.close()
